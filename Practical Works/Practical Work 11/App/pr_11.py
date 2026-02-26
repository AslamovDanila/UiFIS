import flet as ft
import matplotlib.pyplot as plt
from scipy.stats import norm
import numpy as np
import sqlite3
from datetime import datetime
import openpyxl
import io
import base64
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def init_db():
    conn = sqlite3.connect('quality_indexes.db')
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS calculations
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      usl REAL,
                      lsl REAL,
                      mean REAL,
                      std REAL,
                      cp REAL,
                      cpk REAL,
                      status TEXT,
                      timestamp TEXT)''')
    conn.commit()
    conn.close()

def save_calculation(usl, lsl, mean, std, cp, cpk, status):
    conn = sqlite3.connect('quality_indexes.db')
    cursor = conn.cursor()
    cursor.execute('INSERT INTO calculations (usl, lsl, mean, std, cp, cpk, status, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                   (usl, lsl, mean, std, cp, cpk, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()

def load_history():
    conn = sqlite3.connect('quality_indexes.db')
    cursor = conn.cursor()
    cursor.execute('SELECT id, usl, lsl, mean, std, cp, cpk, status, timestamp FROM calculations ORDER BY id DESC')
    rows = cursor.fetchall()
    conn.close()
    return rows

def clear_history():
    conn = sqlite3.connect('quality_indexes.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM calculations')
    conn.commit()
    conn.close()

def export_to_excel(filename, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История расчетов"
    
    headers = ["ID", "Верхняя граница", "Нижняя граница", "Среднее", "Стандартное отклонение", "Cp", "Cpk", "Статус", "Время"]
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1, value=header)
    
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1, value=value)
    
    wb.save(filename)

def import_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                data.append([cell.value for cell in row])
        return data
    except:
        return []

def export_to_pdf(filename, usl, lsl, mean, std, cp, cpk, status):
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
    except:
        pass
    
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    
    style = styles['Normal']
    style.fontName = 'Arial'
    style.leading = 14
    
    style_title = styles['Title']
    style_title.fontName = 'Arial-Bold'
    style_title.alignment = TA_CENTER
    
    style_heading = styles['Heading2']
    style_heading.fontName = 'Arial-Bold'
    
    elements = []
    
    title = Paragraph("Анализ индексов качества процесса", style_title)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("Результаты расчета", style_heading))
    elements.append(Spacer(1, 12))
    
    text = f"""Верхняя граница допуска (USL): {usl:.2f}
Нижняя граница допуска (LSL): {lsl:.2f}
Среднее арифметическое (μ): {mean:.2f}
Стандартное отклонение (σ): {std:.2f}
Индекс воспроизводимости (Cp): {cp:.3f}
Индекс пригодности (Cpk): {cpk:.3f}
Статус: {status}"""
    
    elements.append(Paragraph(text, style))
    elements.append(Spacer(1, 20))
    
    doc.build(elements)

class QualityAnalyzer:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page.title = "Анализ индексов качества процесса"
        self.page.window.width = 1200
        self.page.window.height = 700
        
        init_db()
        
        self.file_picker = ft.FilePicker(on_result=self.pick_files_result)
        self.page.overlay.append(self.file_picker)
        
        self.setup_ui()
    
    def setup_ui(self):
        self.usl_input = ft.TextField(label="Верхняя граница допуска (USL)", width=200)
        self.lsl_input = ft.TextField(label="Нижняя граница допуска (LSL)", width=200)
        self.mean_input = ft.TextField(label="Среднее арифметическое (μ)", width=200)
        self.std_input = ft.TextField(label="Стандартное отклонение (σ)", width=200)
        
        self.cp_result = ft.Text("", size=14, weight=ft.FontWeight.BOLD)
        self.cpk_result = ft.Text("", size=14, weight=ft.FontWeight.BOLD)
        self.status_result = ft.Text("", size=14, weight=ft.FontWeight.BOLD)
        self.action_result = ft.Text("", size=12, color="blue")
        
        self.calculate_btn = ft.ElevatedButton("Рассчитать", on_click=self.calculate)
        self.save_btn = ft.ElevatedButton("Сохранить в историю", on_click=self.save_result, visible=False)
        self.export_excel_btn = ft.ElevatedButton("Экспорт в Excel", on_click=self.export_current, visible=False)
        self.export_pdf_btn = ft.ElevatedButton("Экспорт в PDF", on_click=self.export_to_pdf_current, visible=False)
        self.import_excel_btn = ft.ElevatedButton("Импорт из Excel", on_click=self.import_excel)
        self.open_chart_btn = ft.ElevatedButton("Открыть график", on_click=self.open_chart_window, visible=False)
        
        calculator_content = ft.Container(
            content=ft.Column([
                ft.Row([self.usl_input, self.lsl_input], alignment=ft.MainAxisAlignment.CENTER),
                ft.Row([self.mean_input, self.std_input], alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(height=20),
                ft.Row([self.calculate_btn, self.import_excel_btn], alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(height=20),
                ft.Row([self.save_btn, self.export_excel_btn, self.export_pdf_btn, self.open_chart_btn], alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(height=20),
                ft.Text("Результаты расчета:", size=16, weight=ft.FontWeight.BOLD),
                ft.Container(height=5),
                self.cp_result,
                self.cpk_result,
                self.status_result,
                self.action_result
            ]),
            padding=20
        )
        
        self.chart_image = ft.Image(width=700, height=400)
        refresh_chart_btn = ft.ElevatedButton("Обновить график", on_click=self.refresh_chart)
        chart_content = ft.Container(
            content=ft.Column([
                ft.Text("График распределения", size=18, weight=ft.FontWeight.BOLD),
                ft.Container(height=10),
                self.chart_image,
                ft.Container(height=10),
                refresh_chart_btn,
                ft.Container(height=10),
                ft.Text("Красная зона: вне допуска\nЗеленая зона: в допуске", size=12, color="grey")
            ], alignment=ft.MainAxisAlignment.CENTER),
            padding=20
        )
        
        self.history_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("ID")),
                ft.DataColumn(ft.Text("USL")),
                ft.DataColumn(ft.Text("LSL")),
                ft.DataColumn(ft.Text("μ")),
                ft.DataColumn(ft.Text("σ")),
                ft.DataColumn(ft.Text("Cp")),
                ft.DataColumn(ft.Text("Cpk")),
                ft.DataColumn(ft.Text("Статус")),
                ft.DataColumn(ft.Text("Время"))
            ],
            rows=[],
            height=300
        )
        
        history_content = ft.Container(
            content=ft.Column([
                ft.Text("История расчетов", size=18, weight=ft.FontWeight.BOLD),
                ft.Container(height=10),
                self.history_table,
                ft.Container(height=10),
                ft.Row([
                    ft.ElevatedButton("Очистить историю", on_click=self.clear_history),
                    ft.ElevatedButton("Экспорт истории в Excel", on_click=self.export_history)
                ], alignment=ft.MainAxisAlignment.CENTER)
            ]),
            padding=20
        )
        
        tabs = ft.Tabs(
            selected_index=0,
            tabs=[
                ft.Tab(text="Калькулятор", content=calculator_content),
                ft.Tab(text="График распределения", content=chart_content),
                ft.Tab(text="История расчетов", content=history_content)
            ]
        )
        
        self.page.add(tabs)
        self.update_history_table()
    
    def get_status(self, cpk):
        if cpk >= 1.33:
            return "Отлично"
        elif cpk >= 1.0:
            return "Удовлетворительно"
        elif cpk >= 0.67:
            return "Неудовлетворительно"
        else:
            return "Критично"
    
    def get_actions(self, cpk):
        if cpk >= 1.33:
            return "Процесс пригоден, обычный контроль"
        elif cpk >= 1.0:
            return "Требуется усиленный контроль"
        elif cpk >= 0.67:
            return "Требуется улучшение процесса"
        else:
            return "Процесс непригоден, требуется остановка"
    
    def calculate(self, e):
        try:
            usl = float(self.usl_input.value) if self.usl_input.value else 0
            lsl = float(self.lsl_input.value) if self.lsl_input.value else 0
            mean = float(self.mean_input.value) if self.mean_input.value else 0
            std = float(self.std_input.value) if self.std_input.value else 0
            
            if usl <= lsl:
                self.status_result.value = "Ошибка: верхняя граница должна быть больше нижней!"
                self.status_result.color = "red"
                self.page.update()
                return
            
            if std <= 0:
                self.status_result.value = "Ошибка: стандартное отклонение должно быть положительным!"
                self.status_result.color = "red"
                self.page.update()
                return
            
            cp = (usl - lsl) / (6 * std)
            
            cpu = (usl - mean) / (3 * std)
            cpl = (mean - lsl) / (3 * std)
            cpk = min(cpu, cpl)
            
            status = self.get_status(cpk)
            actions = self.get_actions(cpk)
            
            self.cp_result.value = f"Индекс воспроизводимости (Cp): {cp:.3f}"
            self.cpk_result.value = f"Индекс пригодности (Cpk): {cpk:.3f}"
            self.status_result.value = f"Статус процесса: {status}"
            self.action_result.value = f"Необходимые действия: {actions}"
            
            if cpk >= 1.33:
                self.status_result.color = "green"
                self.action_result.color = "green"
            elif cpk >= 1.0:
                self.status_result.color = "blue"
                self.action_result.color = "blue"
            elif cpk >= 0.67:
                self.status_result.color = "orange"
                self.action_result.color = "orange"
            else:
                self.status_result.color = "red"
                self.action_result.color = "red"
            
            self.save_btn.visible = True
            self.export_excel_btn.visible = True
            self.export_pdf_btn.visible = True
            self.open_chart_btn.visible = True
            
            self.generate_chart(usl, lsl, mean, std)
            
            self.page.update()
            
        except ValueError:
            self.status_result.value = "Ошибка: введите числовые значения!"
            self.status_result.color = "red"
            self.page.update()
    
    def generate_chart(self, usl, lsl, mean, std):
        plt.switch_backend('Agg')
        
        x = np.linspace(lsl - 3 * std, usl + 3 * std, 1000)
        y = norm.pdf(x, mean, std)
        
        fig, ax = plt.subplots(figsize=(8, 5))
        ax.plot(x, y, 'b-', linewidth=2)
        
        ax.fill_between(x, 0, y, where=(x < lsl) | (x > usl), color='red', alpha=0.3, label='Вне допуска')
        ax.fill_between(x, 0, y, where=(x >= lsl) & (x <= usl), color='green', alpha=0.3, label='В допуске')
        
        ax.axvline(x=lsl, color='black', linestyle='--', label=f'LSL = {lsl:.2f}')
        ax.axvline(x=usl, color='black', linestyle='--', label=f'USL = {usl:.2f}')
        ax.axvline(x=mean, color='blue', label=f'μ = {mean:.2f}')
        
        ax.set_xlabel('Значение параметра')
        ax.set_ylabel('Плотность вероятности')
        ax.set_title('Нормальное распределение процесса')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        filename = "chart.png"
        plt.savefig(filename, format='png', dpi=100, bbox_inches='tight')
        plt.close(fig)
        
        self.chart_image.src = filename
        self.chart_image.visible = True
        self.chart_image.update()
        self.page.update()
    
    def save_result(self, e):
        try:
            usl = float(self.usl_input.value) if self.usl_input.value else 0
            lsl = float(self.lsl_input.value) if self.lsl_input.value else 0
            mean = float(self.mean_input.value) if self.mean_input.value else 0
            std = float(self.std_input.value) if self.std_input.value else 0
            
            cp = (usl - lsl) / (6 * std)
            cpu = (usl - mean) / (3 * std)
            cpl = (mean - lsl) / (3 * std)
            cpk = min(cpu, cpl)
            status = self.get_status(cpk)
            
            save_calculation(usl, lsl, mean, std, cp, cpk, status)
            self.update_history_table()
            
            if self.status_result.value:
                self.status_result.value += " | Сохранено!"
            self.page.update()
            
        except:
            self.status_result.value = "Ошибка при сохранении!"
            self.page.update()
    
    def update_history_table(self):
        history = load_history()
        self.history_table.rows = []
        
        for row in history:
            self.history_table.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(row[0]))),
                    ft.DataCell(ft.Text(f"{row[1]:.2f}")),
                    ft.DataCell(ft.Text(f"{row[2]:.2f}")),
                    ft.DataCell(ft.Text(f"{row[3]:.2f}")),
                    ft.DataCell(ft.Text(f"{row[4]:.2f}")),
                    ft.DataCell(ft.Text(f"{row[5]:.3f}")),
                    ft.DataCell(ft.Text(f"{row[6]:.3f}")),
                    ft.DataCell(ft.Text(row[7])),
                    ft.DataCell(ft.Text(row[8]))
                ])
            )
        
        self.page.update()
    
    def clear_history(self, e):
        clear_history()
        self.update_history_table()
    
    def export_current(self, e):
        try:
            usl = float(self.usl_input.value) if self.usl_input.value else 0
            lsl = float(self.lsl_input.value) if self.lsl_input.value else 0
            mean = float(self.mean_input.value) if self.mean_input.value else 0
            std = float(self.std_input.value) if self.std_input.value else 0
            cp = (usl - lsl) / (6 * std)
            cpu = (usl - mean) / (3 * std)
            cpl = (mean - lsl) / (3 * std)
            cpk = min(cpu, cpl)
            status = self.get_status(cpk)
            
            filename = f"пиво_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            data = [[1, usl, lsl, mean, std, cp, cpk, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]]
            export_to_excel(filename, data)
            
            if self.status_result.value:
                self.status_result.value += " | Экспорт в Excel выполнен!"
            else:
                self.status_result.value = "Экспорт в Excel выполнен!"
            self.page.update()
            
        except:
            self.status_result.value = "Ошибка при экспорте!"
            self.page.update()
    
    def export_history(self, e):
        history = load_history()
        if not history:
            self.status_result.value = "История пуста!"
            self.page.update()
            return
        
        filename = f"пиво_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_to_excel(filename, history)
        
        self.status_result.value = "Экспорт истории в Excel выполнен!"
        self.page.update()
    
    def open_chart_window(self, e):
        chart_dialog = ft.AlertDialog(
            title=ft.Text("График распределения"),
            content=ft.Container(
                content=ft.Image(
                    src='chart.png',
                    width=700,
                    height=400
                ),
                padding=10
            ),
            actions=[
                ft.TextButton("Закрыть", on_click=lambda x: self.page.close(chart_dialog))
            ],
            actions_alignment=ft.MainAxisAlignment.END
        )
        self.page.open(chart_dialog)
    
    def refresh_chart(self, e):
        self.chart_image.src = "chart.png"
        self.chart_image.update()
        self.page.update()
    
    def export_to_pdf_current(self, e):
        try:
            usl = float(self.usl_input.value) if self.usl_input.value else 0
            lsl = float(self.lsl_input.value) if self.lsl_input.value else 0
            mean = float(self.mean_input.value) if self.mean_input.value else 0
            std = float(self.std_input.value) if self.std_input.value else 0
            cp = (usl - lsl) / (6 * std)
            cpu = (usl - mean) / (3 * std)
            cpl = (mean - lsl) / (3 * std)
            cpk = min(cpu, cpl)
            status = self.get_status(cpk)
            
            filename = f"пиво_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            export_to_pdf(filename, usl, lsl, mean, std, cp, cpk, status)
            
            if self.status_result.value:
                self.status_result.value += " | Экспорт в PDF выполнен!"
            else:
                self.status_result.value = "Экспорт в PDF выполнен!"
            self.page.update()
            
        except:
            self.status_result.value = "Ошибка при экспорте в PDF!"
            self.page.update()
    
    def import_excel(self, e):
        self.file_picker.pick_files(file_type=ft.FilePickerFileType.CUSTOM, allowed_extensions=["xlsx"])
    
    def pick_files_result(self, e):
        if e.files:
            filename = e.files[0].path
            data = import_from_excel(filename)
            if data:
                row = data[0]
                if len(row) >= 4:
                    try:
                        self.usl_input.value = str(row[1]) if row[1] else ""
                        self.lsl_input.value = str(row[2]) if row[2] else ""
                        self.mean_input.value = str(row[3]) if row[3] else ""
                        self.std_input.value = str(row[4]) if row[4] else ""
                        self.calculate(None)
                    except:
                        pass


def main(page: ft.Page):
    QualityAnalyzer(page)


if __name__ == "__main__":
    init_db()
    ft.app(target=main)
